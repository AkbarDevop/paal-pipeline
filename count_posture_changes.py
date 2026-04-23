#!/usr/bin/env python3
"""Count posture transitions during nighttime hours for the first 10-12 days.

Reads predictions.csv and an estrus Excel file to determine the monitoring
window (Day 0-11). Counts how many times each pig changes posture during
nighttime hours (00:00-08:00 and 19:00-23:59).

Usage:
    python count_posture_changes.py
    python count_posture_changes.py --estrus "February 2026 Estrus.xlsx"
    python count_posture_changes.py --csv outputs/predictions.csv --days 12
    python count_posture_changes.py --night-only false  # include daytime too
"""

import argparse
import csv
import os
from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl

from config import OUTPUT_DIR


POSTURE_NAMES = {0: "standing", 1: "sitting", 2: "lying"}
TRANSITION_TYPES = [
    "standing_to_sitting", "standing_to_lying",
    "sitting_to_standing", "sitting_to_lying",
    "lying_to_standing", "lying_to_sitting",
]


def _parse_date(date_str):
    """Try multiple date formats."""
    for fmt in ("%m/%d/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(date_str), fmt)
        except ValueError:
            continue
    return None


def _find_data_start(ws):
    """Find where Day 0 data starts (row 5 or row 15 depending on format)."""
    for row_idx in range(1, 20):
        val = ws.cell(row=row_idx, column=1).value
        if val == 0 or val == 0.0:
            return row_idx
    return 15  # fallback


def _find_heat_columns(ws, data_start):
    """Find heat morning/evening column indices by scanning headers above data."""
    # Search headers for 'MORNING' and 'EVENING'
    for r in range(max(1, data_start - 3), data_start):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and str(val).upper() == "MORNING":
                return c, c + 1  # MORNING, EVENING are adjacent
    # Fallback: column 16, 17 (Feb format) or 17, 18 (March format)
    return 16, 17


def get_day0_from_estrus(estrus_path):
    """Extract Day 0 date from estrus Excel file (any pig sheet)."""
    wb = openpyxl.load_workbook(estrus_path, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name.isdigit():
            ws = wb[sheet_name]
            start = _find_data_start(ws)
            for row in ws.iter_rows(min_row=start, max_row=start + 40, values_only=True):
                day = row[0]
                date_str = row[1]
                if day == 0 or day == 0.0:
                    if date_str is not None:
                        return _parse_date(date_str)
            break
    return None


def get_estrus_days(estrus_path):
    """Extract heat days per pig from estrus Excel."""
    wb = openpyxl.load_workbook(estrus_path, data_only=True)
    heat_days = defaultdict(set)

    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit():
            continue
        pig_id = int(sheet_name)
        ws = wb[sheet_name]
        data_start = _find_data_start(ws)
        heat_m_col, heat_e_col = _find_heat_columns(ws, data_start)

        for row in ws.iter_rows(min_row=data_start, max_row=data_start + 40, values_only=False):
            date_str = row[1].value
            if date_str is None:
                continue
            heat_m = row[heat_m_col - 1].value if len(row) >= heat_m_col else 0
            heat_e = row[heat_e_col - 1].value if len(row) >= heat_e_col else 0
            heat_m = heat_m if heat_m else 0
            heat_e = heat_e if heat_e else 0
            if heat_m == 20 or heat_e == 20:
                dt = _parse_date(date_str)
                if dt:
                    heat_days[pig_id].add(dt.strftime("%Y%m%d"))
    return heat_days


NIGHT_START_HOUR = 19  # 7 PM
NIGHT_END_HOUR = 8     # 8 AM


def is_nighttime(hour):
    """Check if hour is in nighttime window (19:00-23:59 or 00:00-08:00)."""
    return hour >= NIGHT_START_HOUR or hour < NIGHT_END_HOUR


def night_date(dt):
    """Return the date that this night belongs to (the MORNING date).

    A night runs from 19:00 on day X to 08:00 on day X+1.
    It is attributed to day X+1 (the morning date), matching when Lucas
    runs heat checks in the morning.

    Example:
      2026-02-11 20:00 -> night_date = 2026-02-12
      2026-02-12 02:00 -> night_date = 2026-02-12
      2026-02-12 07:00 -> night_date = 2026-02-12
    """
    if dt.hour >= NIGHT_START_HOUR:
        return (dt + timedelta(days=1)).strftime("%Y%m%d")
    return dt.strftime("%Y%m%d")


def load_predictions(csv_path):
    """Load predictions and parse timestamps."""
    rows = []
    with open(csv_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            ts = row["pig_timestamp"]
            try:
                dt = datetime.strptime(ts, "%Y%m%d-%H-%M-%S")
            except ValueError:
                continue
            rows.append({
                "pig_id": int(row["pig_id"]),
                "datetime": dt,
                "date": dt.strftime("%Y%m%d"),
                "night_date": night_date(dt),
                "hour": dt.hour,
                "posture": row["prediction_name"],
                "confidence": float(row["confidence"]),
            })
    return rows


def count_transitions(predictions, night_only=True, max_gap_minutes=30):
    """Count posture transitions per pig per night (or per date if night_only=False).

    A transition is when consecutive frames (same pig, sorted by time)
    have different postures. Gaps > max_gap_minutes are skipped.
    Posture counts (standing/sitting/lying) are also tracked.

    When night_only=True, each "night" is a single 13-hour session from 19:00
    to 08:00 the next morning, attributed to the morning date.
    """
    # Group by pig_id, sort by time
    by_pig = defaultdict(list)
    for p in predictions:
        if night_only and not is_nighttime(p["hour"]):
            continue
        by_pig[p["pig_id"]].append(p)

    for pid in by_pig:
        by_pig[pid].sort(key=lambda x: x["datetime"])

    # Grouping key: night_date if night_only else date
    key_field = "night_date" if night_only else "date"

    results = defaultdict(lambda: defaultdict(lambda: {
        "total_transitions": 0,
        "total_frames": 0,
        "posture_counts": defaultdict(int),
        "transitions": defaultdict(int),
    }))

    for pid, frames in by_pig.items():
        for i in range(len(frames)):
            group_key = frames[i][key_field]
            results[pid][group_key]["total_frames"] += 1
            results[pid][group_key]["posture_counts"][frames[i]["posture"]] += 1

            if i == 0:
                continue

            prev = frames[i - 1]
            curr = frames[i]
            gap = (curr["datetime"] - prev["datetime"]).total_seconds() / 60

            if gap > max_gap_minutes:
                continue
            # When using night_date, the night crosses calendar dates (19:xx -> 00:xx).
            # We only count transitions within the SAME night_date / date.
            if prev[key_field] != curr[key_field]:
                continue

            if prev["posture"] != curr["posture"]:
                t_type = f"{prev['posture']}_to_{curr['posture']}"
                results[pid][group_key]["total_transitions"] += 1
                results[pid][group_key]["transitions"][t_type] += 1

    return results


def main():
    parser = argparse.ArgumentParser(description="Count posture transitions")
    parser.add_argument("--csv", default=os.path.join(OUTPUT_DIR, "predictions.csv"),
                        help="Path to predictions CSV")
    parser.add_argument("--estrus", default=None,
                        help="Path to estrus Excel file (to get Day 0 and heat days)")
    parser.add_argument("--days", type=int, default=12,
                        help="Number of days to analyze from Day 0 (default: 12)")
    parser.add_argument("--night-only", type=str, default="true",
                        help="Only count nighttime hours (default: true)")
    parser.add_argument("--output", default=None,
                        help="Output CSV path")
    args = parser.parse_args()

    night_only = args.night_only.lower() == "true"

    # Load predictions
    print(f"Loading predictions from {args.csv}...")
    predictions = load_predictions(args.csv)
    print(f"  {len(predictions)} total frames")

    # Determine date range
    all_dates = sorted(set(p["date"] for p in predictions))
    day0 = None
    heat_days = {}

    if args.estrus:
        print(f"Reading estrus file: {args.estrus}")
        day0 = get_day0_from_estrus(args.estrus)
        heat_days = get_estrus_days(args.estrus)
        if day0:
            print(f"  Day 0 = {day0.strftime('%Y-%m-%d')}")
            print(f"  Monitoring window: Day 0-{args.days - 1} ({day0.strftime('%m/%d')} - {(day0 + timedelta(days=args.days - 1)).strftime('%m/%d')})")

    if day0:
        start_date = day0.strftime("%Y%m%d")
        end_date = (day0 + timedelta(days=args.days)).strftime("%Y%m%d")
        predictions = [p for p in predictions if start_date <= p["date"] < end_date]
    else:
        # Use first N days of data
        unique_dates = sorted(set(p["date"] for p in predictions))
        if len(unique_dates) > args.days:
            cutoff = unique_dates[args.days]
            predictions = [p for p in predictions if p["date"] < cutoff]

    dates_used = sorted(set(p["date"] for p in predictions))
    time_label = "nighttime (19:00-08:00 next day)" if night_only else "all hours"
    print(f"  Analyzing {len(predictions)} frames across {len(dates_used)} days ({time_label})")

    # Count transitions
    results = count_transitions(predictions, night_only=night_only)

    # Output CSV
    out_path = args.output or os.path.join(OUTPUT_DIR, "posture_transitions.csv")
    fields = ["pig_id", "date", "day_num", "total_frames", "total_transitions",
              "is_heat_day", "lying", "sitting", "standing",
              "lying_pct", "sitting_pct", "standing_pct"] + TRANSITION_TYPES

    rows = []
    for pid in sorted(results.keys()):
        for date in sorted(results[pid].keys()):
            data = results[pid][date]
            # day_num = days from group's Day 0 (from estrus file)
            if day0:
                date_dt = datetime.strptime(date, "%Y%m%d")
                day_num = (date_dt - day0).days
            else:
                day_num = dates_used.index(date) if date in dates_used else -1
            is_heat = date in heat_days.get(pid, set())
            total_frames = data["total_frames"]
            pc = data["posture_counts"]
            row = {
                "pig_id": pid,
                "date": date,
                "day_num": day_num,
                "total_frames": total_frames,
                "total_transitions": data["total_transitions"],
                "is_heat_day": is_heat,
                "lying": pc.get("lying", 0),
                "sitting": pc.get("sitting", 0),
                "standing": pc.get("standing", 0),
                "lying_pct": round(pc.get("lying", 0) / total_frames * 100, 1) if total_frames else 0,
                "sitting_pct": round(pc.get("sitting", 0) / total_frames * 100, 1) if total_frames else 0,
                "standing_pct": round(pc.get("standing", 0) / total_frames * 100, 1) if total_frames else 0,
            }
            for t in TRANSITION_TYPES:
                row[t] = data["transitions"].get(t, 0)
            rows.append(row)

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nSaved: {out_path}")

    # Print summary
    print(f"\n{'=' * 70}")
    print("POSTURE TRANSITIONS SUMMARY")
    print(f"{'=' * 70}")
    print(f"{'Pig':>5} {'Days':>5} {'Frames':>7} {'Transitions':>12} {'Avg/Night':>10} {'L_to_S':>5} {'S_to_L':>5} {'L_to_Si':>5} {'Si_to_L':>5}")
    print("-" * 70)

    for pid in sorted(results.keys()):
        dates = results[pid]
        total_frames = sum(d["total_frames"] for d in dates.values())
        total_trans = sum(d["total_transitions"] for d in dates.values())
        n_days = len(dates)
        avg = total_trans / n_days if n_days > 0 else 0

        # Aggregate transition types
        agg = defaultdict(int)
        for d in dates.values():
            for t, c in d["transitions"].items():
                agg[t] += c

        print(f"{pid:>5} {n_days:>5} {total_frames:>7} {total_trans:>12} {avg:>10.1f}"
              f" {agg.get('lying_to_standing', 0):>5} {agg.get('standing_to_lying', 0):>5}"
              f" {agg.get('lying_to_sitting', 0):>5} {agg.get('sitting_to_lying', 0):>5}")

    # Heat vs non-heat comparison if estrus data available
    if heat_days:
        print(f"\n{'=' * 70}")
        print("HEAT vs NON-HEAT DAYS (avg transitions per night)")
        print(f"{'=' * 70}")
        print(f"{'Pig':>5} {'Heat Avg':>10} {'Non-Heat Avg':>13} {'Diff':>8}")
        print("-" * 40)

        for pid in sorted(results.keys()):
            heat_trans = []
            noheat_trans = []
            for date, data in results[pid].items():
                if date in heat_days.get(pid, set()):
                    heat_trans.append(data["total_transitions"])
                else:
                    noheat_trans.append(data["total_transitions"])

            if heat_trans:
                h_avg = sum(heat_trans) / len(heat_trans)
                n_avg = sum(noheat_trans) / len(noheat_trans) if noheat_trans else 0
                print(f"{pid:>5} {h_avg:>10.1f} {n_avg:>13.1f} {h_avg - n_avg:>+8.1f}")


if __name__ == "__main__":
    main()
