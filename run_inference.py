#!/usr/bin/env python3
"""One-command posture inference pipeline.

Takes a single input path to raw OAK-D data, crops images, runs the
trained model, and produces a predictions CSV + posture heatmap.

Usage:
    python run_inference.py C:\\PAAL_Data\\fed_pig
    python run_inference.py /path/to/oak_data
"""

import argparse
import csv
import os
import re
import sys
from collections import Counter
from datetime import datetime

import cv2
import numpy as np
import torch

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    HAS_PLT = True
except ImportError:
    HAS_PLT = False

from config import CROP_TOF, MODEL_DIR, OUTPUT_DIR, POSTURE3_CLASSES, IMG_SIZE
from models import SingleModalModel
from data_loader import MODALITY_CHANNELS, load_and_preprocess

# ── Constants ─────────────────────────────────────────────────────────────

TOF_W, TOF_H = 640, 480
DEPTH_THRESHOLD = 1463
BAR_MARGIN = 50

FILENAME_RE = re.compile(
    r"^pig(\d+)_(depth_vis|depth|ir_vis|ir|rgb_aligned|rgb)_"
    r"(\d{8}-\d{2}-\d{2}-\d{2})(_cropped)?\.(jpg|raw)$"
)
FOLDER_RE = re.compile(r"^\d{8}-\d{2}-\d{2}-\d{2}$")


# ── Step 1: Crop images ──────────────────────────────────────────────────

def crop_box_for_size(w, h):
    x1, y1, x2, y2 = CROP_TOF
    sx, sy = w / TOF_W, h / TOF_H
    return int(x1 * sx), int(y1 * sy), int(x2 * sx), int(y2 * sy)


def load_and_crop_ir(ir_path, size=IMG_SIZE):
    """Load raw IR jpg, crop to stall region, resize for model.

    Does everything in-memory — no disk write. Replaces the old
    crop-to-disk step which saved *_cropped.jpg files.
    """
    if not ir_path or not os.path.exists(ir_path):
        return None
    if os.path.getsize(ir_path) < 1000:
        return None  # corrupt
    img = cv2.imread(ir_path)
    if img is None:
        return None
    h, w = img.shape[:2]
    if h < 10 or w < 10:
        return None
    x1, y1, x2, y2 = crop_box_for_size(w, h)
    cropped = img[y1:y2, x1:x2]
    cropped = cv2.resize(cropped, (size, size))
    cropped = cv2.cvtColor(cropped, cv2.COLOR_BGR2RGB).astype(np.float32) / 255.0
    return cropped


# ── Step 2: Scan + prefilter + predict ────────────────────────────────────

def load_depth_raw(path):
    if not path or not os.path.exists(path):
        return None
    size = os.path.getsize(path)
    expected = TOF_W * TOF_H * 2
    if size == expected + 8:
        raw = np.fromfile(path, dtype=np.uint16, offset=8)
    elif size == expected:
        raw = np.fromfile(path, dtype=np.uint16)
    else:
        return None
    if raw.size != TOF_W * TOF_H:
        return None
    return raw.reshape((TOF_H, TOF_W))


def check_pig_present(depth_raw_path):
    depth = load_depth_raw(depth_raw_path)
    if depth is None:
        return True, 0.0
    x1, y1, x2, y2 = CROP_TOF
    crop = depth[y1:y2, x1:x2].copy()
    crop[:, :BAR_MARGIN] = 0
    crop[:, -BAR_MARGIN:] = 0
    valid = crop[(crop > 200) & (crop < 5000)]
    if len(valid) == 0:
        return False, 0.0
    median = float(np.median(valid))
    return median < DEPTH_THRESHOLD, median


def scan_folder(data_dir):
    folders = sorted(
        d for d in os.listdir(data_dir)
        if os.path.isdir(os.path.join(data_dir, d)) and FOLDER_RE.match(d)
    )
    if not folders:
        return []

    records = {}
    for folder in folders:
        folder_path = os.path.join(data_dir, folder)
        for fname in os.listdir(folder_path):
            m = FILENAME_RE.match(fname)
            if not m:
                continue
            pig_id = int(m.group(1))
            modality = m.group(2)
            timestamp = m.group(3)
            is_cropped = m.group(4) is not None
            ext = m.group(5)

            key = (folder, pig_id, timestamp)
            if key not in records:
                records[key] = {
                    "timestamp_folder": folder,
                    "pig_id": pig_id,
                    "pig_timestamp": timestamp,
                }

            if modality == "ir_vis":
                mod_key = "ir"
            elif modality == "depth_vis":
                mod_key = "depth"
            else:
                mod_key = modality

            col_name = f"{mod_key}{'_cropped' if is_cropped else ''}_{ext}"
            records[key][col_name] = os.path.join(folder_path, fname)

    return sorted(records.values(), key=lambda r: (r["timestamp_folder"], r["pig_id"]))


def get_image_path(frame):
    """Get uncropped IR path (cropping now happens in-memory)."""
    p = frame.get("ir_jpg", "")
    if p and os.path.exists(p):
        return p
    return ""


def load_pig_detector(model_dir, device):
    """Load the binary pig-presence CNN. Returns None if weights not found."""
    path = os.path.join(model_dir, "pig_detector.pth")
    if not os.path.exists(path):
        return None
    import torchvision.models as tv_models
    detector = tv_models.mobilenet_v2(weights=None)
    detector.classifier[1] = torch.nn.Linear(detector.classifier[1].in_features, 2)
    ckpt = torch.load(path, map_location=device, weights_only=False)
    detector.load_state_dict(ckpt["model_state_dict"])
    detector.eval().to(device)
    return detector


def run_predictions(data_dir, device, model, detector=None):
    frames = scan_folder(data_dir)
    if not frames:
        print("No frames found.")
        return []

    print(f"[1/2] Running inference on {len(frames)} frames...")
    if detector is None:
        print("  (pig presence: depth prefilter only — detector model not loaded)")
    else:
        print("  (pig presence: depth prefilter + CNN detector)")

    results = []
    skipped_depth = 0
    skipped_detector = 0
    skipped_corrupt = 0

    with torch.no_grad():
        for i, frame in enumerate(frames):
            # Step 1: Depth prefilter (cheap, rejects obvious empty stalls)
            depth_raw = frame.get("depth_raw", "")
            present, median_d = check_pig_present(depth_raw)
            if not present:
                skipped_depth += 1
                continue

            # Step 2: Load + crop IR image
            path = get_image_path(frame)
            if not path:
                skipped_corrupt += 1
                continue
            img = load_and_crop_ir(path)
            if img is None:
                skipped_corrupt += 1
                continue

            x = torch.from_numpy(img).permute(2, 0, 1).unsqueeze(0).to(device)

            # Step 3: CNN pig-presence detector (catches what depth prefilter misses)
            if detector is not None:
                det_logits = detector(x)
                det_probs = torch.softmax(det_logits, dim=1).cpu().numpy()[0]
                # class 0 = empty, class 1 = present
                if det_probs[1] < 0.5:
                    skipped_detector += 1
                    continue

            # Step 4: Posture classification
            logits = model(x)
            probs = torch.softmax(logits, dim=1).cpu().numpy()[0]
            pred = int(np.argmax(probs))
            conf = float(probs[pred])

            results.append({
                "timestamp_folder": frame["timestamp_folder"],
                "pig_id": frame["pig_id"],
                "pig_timestamp": frame["pig_timestamp"],
                "prediction": pred,
                "prediction_name": POSTURE3_CLASSES.get(pred, str(pred)),
                "confidence": round(conf, 4),
                "median_depth": round(median_d, 1),
                "image_path": path,
            })

            if (i + 1) % 500 == 0:
                print(f"  {i + 1}/{len(frames)} frames...")

    print(f"  Predicted: {len(results)}, "
          f"Skipped (depth prefilter): {skipped_depth}, "
          f"Skipped (CNN detector): {skipped_detector}, "
          f"Skipped (corrupt/missing): {skipped_corrupt}")
    return results


# ── Step 3: Generate outputs ──────────────────────────────────────────────

def save_csv(results, out_path):
    fields = ["timestamp_folder", "pig_id", "pig_timestamp",
              "prediction", "prediction_name", "confidence",
              "median_depth", "image_path"]
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(results)


def parse_timestamp(ts_str):
    """Parse pig_timestamp like '20260211-09-17-49' to datetime."""
    try:
        return datetime.strptime(ts_str, "%Y%m%d-%H-%M-%S")
    except ValueError:
        return None


def normalize_pig_id(pid, num_stalls=20):
    """Camera counter can overflow past num_stalls. Map back with modulo."""
    return pid % num_stalls


def generate_heatmap(results, out_path):
    """Generate a posture timeline heatmap (time x pig_id)."""
    if not HAS_PLT:
        print("  matplotlib not installed, skipping heatmap")
        return

    # Normalize pig IDs (camera overflow: 20→0, 21→1, etc.)
    for r in results:
        r["pig_id"] = normalize_pig_id(r["pig_id"])

    # Fixed Y-axis: always pig 0–19 regardless of which pigs have data
    pig_ids = list(range(20))
    timestamps = [parse_timestamp(r["pig_timestamp"]) for r in results]
    valid = [(r, t) for r, t in zip(results, timestamps) if t is not None]
    if not valid:
        print("  No valid timestamps for heatmap")
        return

    min_t = min(t for _, t in valid)
    max_t = max(t for _, t in valid)

    # Create hour-level bins
    from datetime import timedelta
    hours = []
    t = min_t.replace(minute=0, second=0)
    while t <= max_t:
        hours.append(t)
        t += timedelta(hours=1)

    if not hours or not pig_ids:
        return

    # Posture encoding: standing=0, sitting=1, lying=2, no_data=-1
    posture_map = {"standing": 0, "sitting": 1, "lying": 2}

    # Majority vote per (pig, hour) cell
    from collections import defaultdict
    votes = defaultdict(list)
    for r, t in valid:
        if t is None:
            continue
        hi = int((t - hours[0]).total_seconds() / 3600)
        hi = min(hi, len(hours) - 1)
        pid = r["pig_id"]
        posture = posture_map.get(r["prediction_name"], -1)
        votes[(pid, hi)].append(posture)

    pig_idx = {pid: i for i, pid in enumerate(pig_ids)}
    grid = np.full((len(pig_ids), len(hours)), -1, dtype=float)
    for (pid, hi), postures in votes.items():
        # Majority vote: most common posture in this hour
        grid[pig_idx[pid], hi] = max(set(postures), key=postures.count)

    # Custom colormap: gray=no_data, green=standing, orange=sitting, blue=lying
    from matplotlib.colors import ListedColormap, BoundaryNorm
    colors = ["#d4d4d4", "#22c55e", "#f97316", "#3b82f6"]
    cmap = ListedColormap(colors)
    bounds = [-1.5, -0.5, 0.5, 1.5, 2.5]
    norm = BoundaryNorm(bounds, cmap.N)

    fig, ax = plt.subplots(figsize=(max(14, len(hours) * 0.3), 8))
    im = ax.pcolormesh(grid, cmap=cmap, norm=norm, edgecolors="white", linewidth=0.5)

    # Y-axis: fixed pig 0–19
    ax.set_yticks(np.arange(len(pig_ids)) + 0.5)
    ax.set_yticklabels([f"pig {pid}" for pid in pig_ids], fontsize=8)
    ax.set_ylim(0, len(pig_ids))

    # X-axis: fixed 6-hour tick intervals
    step = max(1, 6)
    ax.set_xticks(np.arange(0, len(hours), step) + 0.5)
    ax.set_xticklabels(
        [hours[i].strftime("%m/%d %H:%M") for i in range(0, len(hours), step)],
        rotation=45, ha="right", fontsize=7,
    )
    ax.set_xlim(0, len(hours))

    ax.set_xlabel("Time")
    ax.set_ylabel("Pig ID")
    ax.set_title("Posture Timeline Heatmap (hourly majority vote)")

    # Legend
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor="#22c55e", label="Standing"),
        Patch(facecolor="#f97316", label="Sitting"),
        Patch(facecolor="#3b82f6", label="Lying"),
        Patch(facecolor="#d4d4d4", label="No data"),
    ]
    ax.legend(handles=legend_elements, loc="upper right", fontsize=8)

    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()


def generate_heatmap_raw(results, out_path):
    """Generate heatmap with RAW pig IDs (including overflow 20+).

    This shows all pig IDs the camera assigned, without normalization,
    so you can visually spot which timestamps have duplicates (pig20, pig21, etc).
    """
    if not HAS_PLT:
        print("  matplotlib not installed, skipping raw heatmap")
        return

    from datetime import timedelta
    from collections import defaultdict
    from matplotlib.colors import ListedColormap, BoundaryNorm
    from matplotlib.patches import Patch

    posture_map = {"standing": 0, "sitting": 1, "lying": 2}

    # Use ALL pig IDs found in data (including overflow)
    all_pig_ids = sorted(set(r["pig_id"] for r in results))
    max_pid = max(all_pig_ids) if all_pig_ids else 19

    # Y-axis: pig 0 through max pig ID
    pig_ids = list(range(max_pid + 1))

    timestamps = [parse_timestamp(r["pig_timestamp"]) for r in results]
    valid = [(r, t) for r, t in zip(results, timestamps) if t is not None]
    if not valid:
        return

    min_t = min(t for _, t in valid)
    max_t = max(t for _, t in valid)
    hours = []
    t = min_t.replace(minute=0, second=0)
    while t <= max_t:
        hours.append(t)
        t += timedelta(hours=1)

    if not hours or not pig_ids:
        return

    votes = defaultdict(list)
    for r, t in valid:
        hi = int((t - hours[0]).total_seconds() / 3600)
        hi = min(hi, len(hours) - 1)
        pid = r["pig_id"]
        posture = posture_map.get(r["prediction_name"], -1)
        votes[(pid, hi)].append(posture)

    pig_idx = {pid: i for i, pid in enumerate(pig_ids)}
    grid = np.full((len(pig_ids), len(hours)), -1, dtype=float)
    for (pid, hi), postures in votes.items():
        if pid in pig_idx:
            grid[pig_idx[pid], hi] = max(set(postures), key=postures.count)

    colors = ["#d4d4d4", "#22c55e", "#f97316", "#3b82f6"]
    cmap = ListedColormap(colors)
    bounds = [-1.5, -0.5, 0.5, 1.5, 2.5]
    norm = BoundaryNorm(bounds, cmap.N)

    fig, ax = plt.subplots(figsize=(max(14, len(hours) * 0.3), max(8, len(pig_ids) * 0.4)))
    ax.pcolormesh(grid, cmap=cmap, norm=norm, edgecolors="white", linewidth=0.5)

    ax.set_yticks(np.arange(len(pig_ids)) + 0.5)
    labels = []
    for pid in pig_ids:
        if pid >= 20:
            labels.append(f"pig {pid} (overflow)")
        else:
            labels.append(f"pig {pid}")
    ax.set_yticklabels(labels, fontsize=7)
    ax.set_ylim(0, len(pig_ids))

    # Draw a red line at pig 20 to separate normal from overflow
    if max_pid >= 20:
        ax.axhline(y=20, color="red", linewidth=2, linestyle="--")

    step = max(1, 6)
    ax.set_xticks(np.arange(0, len(hours), step) + 0.5)
    ax.set_xticklabels(
        [hours[i].strftime("%m/%d %H:%M") for i in range(0, len(hours), step)],
        rotation=45, ha="right", fontsize=7,
    )
    ax.set_xlim(0, len(hours))

    ax.set_xlabel("Time")
    ax.set_ylabel("Camera Pig ID (raw, unnormalized)")
    ax.set_title("Posture Heatmap — RAW Pig IDs (overflow visible above red line)")

    legend_elements = [
        Patch(facecolor="#22c55e", label="Standing"),
        Patch(facecolor="#f97316", label="Sitting"),
        Patch(facecolor="#3b82f6", label="Lying"),
        Patch(facecolor="#d4d4d4", label="No data"),
    ]
    ax.legend(handles=legend_elements, loc="upper right", fontsize=8)

    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()
    print(f"  Raw ID heatmap: {out_path}")


def generate_per_pig_heatmaps(results, out_dir):
    """Generate individual heatmap per pig ID."""
    if not HAS_PLT:
        return

    from datetime import timedelta
    from matplotlib.colors import ListedColormap, BoundaryNorm
    from matplotlib.patches import Patch

    posture_map = {"standing": 0, "sitting": 1, "lying": 2}
    colors = ["#d4d4d4", "#22c55e", "#f97316", "#3b82f6"]
    cmap = ListedColormap(colors)
    bounds = [-1.5, -0.5, 0.5, 1.5, 2.5]
    norm = BoundaryNorm(bounds, cmap.N)

    pig_ids = sorted(set(r["pig_id"] for r in results))
    timestamps = [(r, parse_timestamp(r["pig_timestamp"])) for r in results]
    valid = [(r, t) for r, t in timestamps if t is not None]
    if not valid:
        return

    min_t = min(t for _, t in valid)
    max_t = max(t for _, t in valid)
    hours = []
    t = min_t.replace(minute=0, second=0)
    while t <= max_t:
        hours.append(t)
        t += timedelta(hours=1)

    pig_dir = os.path.join(out_dir, "per_pig")
    os.makedirs(pig_dir, exist_ok=True)

    for pid in pig_ids:
        pig_results = [(r, t) for r, t in valid if r["pig_id"] == pid]
        if not pig_results:
            continue

        from collections import defaultdict
        votes = defaultdict(list)
        for r, t in pig_results:
            hi = int((t - hours[0]).total_seconds() / 3600)
            hi = min(hi, len(hours) - 1)
            votes[hi].append(posture_map.get(r["prediction_name"], -1))

        grid = np.full((1, len(hours)), -1, dtype=float)
        for hi, postures in votes.items():
            grid[0, hi] = max(set(postures), key=postures.count)

        _, ax = plt.subplots(figsize=(max(14, len(hours) * 0.15), 2))
        ax.pcolormesh(grid, cmap=cmap, norm=norm, edgecolors="white", linewidth=0.5)
        step = max(1, 6)
        ax.set_xticks(np.arange(0, len(hours), step) + 0.5)
        ax.set_xticklabels(
            [hours[i].strftime("%m/%d %H:%M") for i in range(0, len(hours), step)],
            rotation=45, ha="right", fontsize=7,
        )
        ax.set_xlim(0, len(hours))
        ax.set_yticks([0.5])
        ax.set_yticklabels([f"pig {pid}"], fontsize=9)
        ax.set_title(f"Pig {pid} — Posture Timeline", fontsize=11)

        legend_elements = [
            Patch(facecolor="#22c55e", label="Standing"),
            Patch(facecolor="#f97316", label="Sitting"),
            Patch(facecolor="#3b82f6", label="Lying"),
            Patch(facecolor="#d4d4d4", label="No data"),
        ]
        ax.legend(handles=legend_elements, loc="upper right", fontsize=7, ncol=4)
        plt.tight_layout()
        plt.savefig(os.path.join(pig_dir, f"pig{pid}_heatmap.png"), dpi=150)
        plt.close()

    print(f"  Per-pig heatmaps: {pig_dir}/ ({len(pig_ids)} pigs)")


def generate_per_pig_csvs(results, out_dir):
    """Generate a separate CSV file for each pig ID."""
    pig_dir = os.path.join(out_dir, "per_pig")
    os.makedirs(pig_dir, exist_ok=True)

    fields = ["timestamp_folder", "pig_id", "pig_timestamp",
              "prediction", "prediction_name", "confidence",
              "median_depth", "image_path"]

    pig_ids = sorted(set(r["pig_id"] for r in results))
    for pid in pig_ids:
        pig_results = [r for r in results if r["pig_id"] == pid]
        csv_path = os.path.join(pig_dir, f"pig{pid}_predictions.csv")
        with open(csv_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(pig_results)

    print(f"  Per-pig CSVs: {pig_dir}/ ({len(pig_ids)} pigs)")


def generate_per_pig_excels(results, out_dir):
    """Generate a separate Excel file for each pig ID."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  openpyxl not installed, skipping per-pig Excel files")
        return

    pig_dir = os.path.join(out_dir, "per_pig")
    os.makedirs(pig_dir, exist_ok=True)

    fill_standing = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_sitting  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_lying    = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_header   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font_header   = Font(bold=True, color="FFFFFF", size=11)
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    posture_fills = {"standing": fill_standing, "sitting": fill_sitting, "lying": fill_lying}

    pig_ids = sorted(set(r["pig_id"] for r in results))

    for pid in pig_ids:
        pig_results = sorted(
            [r for r in results if r["pig_id"] == pid],
            key=lambda r: r["pig_timestamp"],
        )
        if not pig_results:
            continue

        wb = Workbook()
        ws = wb.active
        ws.title = f"Pig {pid}"

        headers = ["Date", "Time", "Posture", "Confidence %", "Median Depth (mm)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for i, r in enumerate(pig_results, start=2):
            t = parse_timestamp(r["pig_timestamp"])
            ws.cell(row=i, column=1, value=t.strftime("%Y-%m-%d") if t else "")
            ws.cell(row=i, column=2, value=t.strftime("%H:%M:%S") if t else "")
            posture_cell = ws.cell(row=i, column=3, value=r["prediction_name"].capitalize())
            ws.cell(row=i, column=4, value=round(r["confidence"] * 100, 1))
            ws.cell(row=i, column=5, value=r["median_depth"])

            fill = posture_fills.get(r["prediction_name"])
            if fill:
                posture_cell.fill = fill

            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).border = thin_border
                ws.cell(row=i, column=col).alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 20)

        wb.save(os.path.join(pig_dir, f"pig{pid}_report.xlsx"))

    print(f"  Per-pig Excel files: {pig_dir}/ ({len(pig_ids)} pigs)")


def generate_excel_report(results, out_path):
    """Generate an Excel workbook with summary + per-pig sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  openpyxl not installed, skipping Excel report (pip install openpyxl)")
        return

    from collections import defaultdict

    wb = Workbook()

    # ── Color fills for postures ──
    fill_standing = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
    fill_sitting  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # orange
    fill_lying    = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # blue
    fill_header   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # dark blue
    font_header   = Font(bold=True, color="FFFFFF", size=11)
    font_bold     = Font(bold=True, size=11)
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    posture_fills = {"standing": fill_standing, "sitting": fill_sitting, "lying": fill_lying}

    def style_header(ws, row, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # ── Sheet 1: Summary (daily stats per pig) ──
    ws = wb.active
    ws.title = "Summary"

    # Build daily stats: {pig_id: {date_str: {standing, sitting, lying, total}}}
    pig_ids = sorted(set(r["pig_id"] for r in results))
    daily = defaultdict(lambda: defaultdict(lambda: {"standing": 0, "sitting": 0, "lying": 0, "total": 0}))
    all_dates = set()
    for r in results:
        t = parse_timestamp(r["pig_timestamp"])
        if t is None:
            continue
        date_str = t.strftime("%Y-%m-%d")
        all_dates.add(date_str)
        daily[r["pig_id"]][date_str][r["prediction_name"]] += 1
        daily[r["pig_id"]][date_str]["total"] += 1
    all_dates = sorted(all_dates)

    # Header
    headers = ["Pig ID", "Date", "Total Obs", "Standing", "Sitting", "Lying",
               "Standing %", "Sitting %", "Lying %"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for pid in pig_ids:
        for date_str in all_dates:
            s = daily[pid][date_str]
            if s["total"] == 0:
                continue
            t = s["total"]
            ws.cell(row=row, column=1, value=f"Pig {pid}").font = font_bold
            ws.cell(row=row, column=2, value=date_str)
            ws.cell(row=row, column=3, value=t)
            ws.cell(row=row, column=4, value=s["standing"])
            ws.cell(row=row, column=5, value=s["sitting"])
            ws.cell(row=row, column=6, value=s["lying"])
            ws.cell(row=row, column=7, value=round(s["standing"] / t * 100, 1))
            ws.cell(row=row, column=8, value=round(s["sitting"] / t * 100, 1))
            ws.cell(row=row, column=9, value=round(s["lying"] / t * 100, 1))
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            row += 1

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 20)

    # ── Per-pig sheets ──
    for pid in pig_ids:
        pig_results = sorted(
            [r for r in results if r["pig_id"] == pid],
            key=lambda r: r["pig_timestamp"],
        )
        if not pig_results:
            continue

        ws = wb.create_sheet(title=f"Pig {pid}")
        headers = ["Date", "Time", "Posture", "Confidence %", "Median Depth (mm)"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header(ws, 1, len(headers))

        for i, r in enumerate(pig_results, start=2):
            t = parse_timestamp(r["pig_timestamp"])
            ws.cell(row=i, column=1, value=t.strftime("%Y-%m-%d") if t else "")
            ws.cell(row=i, column=2, value=t.strftime("%H:%M:%S") if t else "")
            posture_cell = ws.cell(row=i, column=3, value=r["prediction_name"].capitalize())
            ws.cell(row=i, column=4, value=round(r["confidence"] * 100, 1))
            ws.cell(row=i, column=5, value=r["median_depth"])

            # Color-code posture cell
            fill = posture_fills.get(r["prediction_name"])
            if fill:
                posture_cell.fill = fill

            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).border = thin_border
                ws.cell(row=i, column=col).alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 20)

    wb.save(out_path)
    print(f"  Excel report: {out_path}")


def generate_timestamp_summary(results, out_path):
    """Generate per-timestamp-folder summary CSV."""
    from collections import defaultdict
    folder_stats = defaultdict(lambda: {"standing": 0, "sitting": 0, "lying": 0, "total": 0, "conf_sum": 0.0})

    for r in results:
        folder = r["timestamp_folder"]
        name = r["prediction_name"]
        folder_stats[folder][name] += 1
        folder_stats[folder]["total"] += 1
        folder_stats[folder]["conf_sum"] += r["confidence"]

    fields = ["timestamp_folder", "total_frames", "standing", "sitting", "lying",
              "standing_pct", "sitting_pct", "lying_pct", "avg_confidence"]

    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for folder in sorted(folder_stats):
            s = folder_stats[folder]
            t = s["total"]
            writer.writerow({
                "timestamp_folder": folder,
                "total_frames": t,
                "standing": s["standing"],
                "sitting": s["sitting"],
                "lying": s["lying"],
                "standing_pct": round(s["standing"] / t * 100, 1),
                "sitting_pct": round(s["sitting"] / t * 100, 1),
                "lying_pct": round(s["lying"] / t * 100, 1),
                "avg_confidence": round(s["conf_sum"] / t, 4),
            })

    print(f"  Timestamp summary: {out_path}")


def print_summary(results):
    counts = Counter(r["prediction_name"] for r in results)
    total = len(results)
    print(f"\n{'='*50}")
    print(f"  Total predicted: {total}")
    for name in ["standing", "sitting", "lying"]:
        c = counts.get(name, 0)
        print(f"  {name:<10}: {c:>6} ({c/total*100:.1f}%)")
    print(f"  Avg confidence: {np.mean([r['confidence'] for r in results]):.3f}")
    print(f"{'='*50}")


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="One-command posture inference: input path → CSV + heatmap"
    )
    parser.add_argument("data_dir", help="Path to OAK-D data folder")
    parser.add_argument("--output-dir", default=None,
                        help="Output directory (default: outputs/)")
    args = parser.parse_args()

    data_dir = args.data_dir
    out_dir = args.output_dir or OUTPUT_DIR

    if not os.path.isdir(data_dir):
        print(f"Error: {data_dir} is not a directory")
        sys.exit(1)

    os.makedirs(out_dir, exist_ok=True)

    # Device
    device = torch.device(
        "mps" if torch.backends.mps.is_available()
        else "cuda" if torch.cuda.is_available()
        else "cpu"
    )
    print(f"Device: {device}")

    # Load model
    model_path = os.path.join(MODEL_DIR, "posture3_ir_best.pth")
    if not os.path.exists(model_path):
        print(f"Error: model not found at {model_path}")
        sys.exit(1)

    ckpt = torch.load(model_path, map_location=device, weights_only=False)
    backbone = ckpt.get("backbone", "mobilenet_v2")
    model = SingleModalModel(
        in_channels=ckpt.get("in_channels", 3),
        num_classes=ckpt.get("num_classes", 3),
        pretrained=False,
        backbone=backbone,
    ).to(device)
    model.load_state_dict(ckpt["model_state_dict"])
    model.eval()
    print(f"Posture model: {backbone}, {ckpt.get('num_classes', 3)} classes")

    # Load pig presence detector (binary CNN)
    detector = load_pig_detector(MODEL_DIR, device)
    if detector is not None:
        print("Pig detector: MobileNetV2 (binary, present/empty)")
    else:
        print("Pig detector: NOT FOUND — falling back to depth prefilter only")
    print()

    # Step 1: Predict (cropping happens in-memory, no disk writes)
    results = run_predictions(data_dir, device, model, detector=detector)
    if not results:
        print("No predictions generated.")
        return

    # Step 2: Save outputs
    csv_path = os.path.join(out_dir, "predictions.csv")
    heatmap_path = os.path.join(out_dir, "posture_heatmap.png")
    heatmap_raw_path = os.path.join(out_dir, "posture_heatmap_raw_ids.png")
    ts_summary_path = os.path.join(out_dir, "timestamp_summary.csv")

    # Save CSV with raw (unnormalized) pig IDs
    save_csv(results, csv_path)
    print(f"\n[2/2] Generating outputs...")

    # Generate heatmap with RAW pig IDs (including overflow 20+)
    generate_heatmap_raw(results, heatmap_raw_path)

    # Normalize pig IDs for the standard heatmap and per-pig outputs
    for r in results:
        r["pig_id"] = normalize_pig_id(r["pig_id"])

    generate_heatmap(results, heatmap_path)
    generate_per_pig_heatmaps(results, out_dir)
    generate_per_pig_csvs(results, out_dir)
    generate_per_pig_excels(results, out_dir)
    excel_path = os.path.join(out_dir, "posture_report.xlsx")
    generate_excel_report(results, excel_path)
    generate_timestamp_summary(results, ts_summary_path)

    print_summary(results)
    print(f"\n  CSV:              {csv_path}")
    print(f"  Heatmap (all):    {heatmap_path}")
    print(f"  Heatmap (raw IDs):{heatmap_raw_path}")
    print(f"  Per-pig outputs:  {os.path.join(out_dir, 'per_pig')}/")
    print(f"  Excel report:     {excel_path}")
    print(f"  Timestamp summary:{ts_summary_path}")


if __name__ == "__main__":
    main()
