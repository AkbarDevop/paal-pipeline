#!/usr/bin/env python3
"""Generate ONE master Excel report per group combining:
  - Overview sheet: summary table + group-wide plots (all pigs transitions, posture breakdown)
  - Per-pig sheets (P0-P19):
    * Daily summary table (nights)
    * Embedded individual plot (transitions + posture stacked-area)
    * Raw observations table with color-coded postures + "change" column (0/1)

Replaces both posture_report.xlsx and per_pig_summary_*.xlsx.

Usage:
    python generate_master_report.py
"""

import csv
import os
import shutil
from collections import defaultdict
from datetime import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR

GROUPS = [
    {
        "name": "Feb Group (020626)",
        "short": "feb",
        "transitions_csv": os.path.join(OUTPUT_DIR, "posture_transitions_feb_all.csv"),
        "predictions_csv": os.path.join(OUTPUT_DIR, "predictions_feb_filtered.csv"),
        "estrus": r"C:\Users\aakdbx\Downloads\February 2026 Estrus.xlsx",
        "sow_source": "title",
    },
    {
        "name": "March Group (031326)",
        "short": "march",
        "transitions_csv": os.path.join(OUTPUT_DIR, "posture_transitions_march_all.csv"),
        "predictions_csv": os.path.join(OUTPUT_DIR, "predictions_march_clean.csv"),
        "estrus": r"C:\Users\aakdbx\Downloads\March 2026 Estrus.xlsx",
        "sow_source": "column",
    },
]

POSTURE_COLORS = {
    "standing": "22c55e",  # green
    "sitting": "f97316",   # orange
    "lying": "3b82f6",     # blue
}


def load_rows(csv_path):
    with open(csv_path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def get_stall_to_sow(estrus_path, source):
    wb = openpyxl.load_workbook(estrus_path, data_only=True)
    mapping = {}
    for stall in range(20):
        sname = str(stall)
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        if source == "title":
            title = ws.cell(row=1, column=1).value
            mapping[stall] = str(title).strip() if title and "SOW" in str(title).upper() else f"Stall {stall}"
        else:
            for r in range(4, 8):
                v = ws.cell(row=r, column=3).value
                if v and str(v) != "X":
                    try:
                        mapping[stall] = f"Tag {int(v)}"
                    except (ValueError, TypeError):
                        mapping[stall] = f"Tag {v}"
                    break
            else:
                mapping[stall] = f"Stall {stall}"
    return mapping


def _save_tmp(fig, key):
    tmp = os.path.join(OUTPUT_DIR, ".plot_tmp")
    os.makedirs(tmp, exist_ok=True)
    path = os.path.join(tmp, f"{key}.png")
    fig.savefig(path, dpi=110, bbox_inches="tight")
    plt.close(fig)
    return path


def plot_all_pigs_transitions(rows, stall_to_sow, title, key):
    fig, ax = plt.subplots(figsize=(14, 7))
    by_pig = defaultdict(list)
    heat_by_pig = defaultdict(set)
    for r in rows:
        pid = int(r["pig_id"])
        by_pig[pid].append((int(r["day_num"]), int(r["total_transitions"])))
        if r["is_heat_day"] == "True":
            heat_by_pig[pid].add(int(r["day_num"]))
    cmap = plt.get_cmap("tab20")
    for i, pid in enumerate(sorted(by_pig.keys())):
        data = sorted(by_pig[pid])
        days = [d for d, _ in data]
        vals = [v for _, v in data]
        label = f"P{pid} / {stall_to_sow.get(pid, f'Stall {pid}')}"
        color = cmap(i % 20)
        ax.plot(days, vals, marker="o", markersize=3, linewidth=1.3, color=color, label=label, alpha=0.85)
        heat_pts = [(d, v) for d, v in zip(days, vals) if d in heat_by_pig[pid]]
        if heat_pts:
            hx, hy = zip(*heat_pts)
            ax.scatter(hx, hy, marker="*", s=150, color=color, edgecolor="black", linewidth=0.6, zorder=5)
    ax.set_xlabel("Day # (from group's Day 0)")
    ax.set_ylabel("Posture transitions per night")
    ax.set_title(f"{title}\nAll Pigs — Nightly Transitions (19:00-08:00). Stars = heat day")
    ax.grid(True, alpha=0.3)
    ax.legend(loc="center left", bbox_to_anchor=(1.01, 0.5), fontsize=7)
    plt.tight_layout()
    return _save_tmp(fig, key)


def plot_all_pigs_postures(rows, stall_to_sow, title, key):
    pigs = sorted(set(int(r["pig_id"]) for r in rows))
    n = len(pigs); cols = 4
    rows_n = (n + cols - 1) // cols
    fig, axes = plt.subplots(rows_n, cols, figsize=(18, 3.2 * rows_n), sharex=True)
    axes = axes.flatten()
    for i, pid in enumerate(pigs):
        ax = axes[i]
        pig_rows = sorted([r for r in rows if int(r["pig_id"]) == pid], key=lambda x: int(x["day_num"]))
        days = [int(r["day_num"]) for r in pig_rows]
        ly = [float(r.get("lying_pct", 0) or 0) for r in pig_rows]
        si = [float(r.get("sitting_pct", 0) or 0) for r in pig_rows]
        st = [float(r.get("standing_pct", 0) or 0) for r in pig_rows]
        heat = [r["is_heat_day"] == "True" for r in pig_rows]
        ax.stackplot(days, ly, si, st, labels=["Lying", "Sitting", "Standing"],
                     colors=["#3b82f6", "#f97316", "#22c55e"], alpha=0.85)
        for d, h in zip(days, heat):
            if h:
                ax.axvline(d, color="red", linestyle="--", linewidth=1, alpha=0.7)
        sow = stall_to_sow.get(pid, f"Stall {pid}")
        ax.set_title(f"P{pid} / {sow}", fontsize=9)
        ax.set_ylim(0, 100); ax.grid(True, alpha=0.3)
        if i == 0:
            ax.legend(loc="upper right", fontsize=6)
    for j in range(len(pigs), len(axes)):
        axes[j].axis("off")
    fig.suptitle(f"{title}\nPer-pig posture % over time (red dashed = heat day)", fontsize=13, y=1.00)
    fig.supxlabel("Day # (from group's Day 0)")
    fig.supylabel("Posture %")
    plt.tight_layout()
    return _save_tmp(fig, key)


def plot_individual_pig(rows, pid, stall_to_sow, group_name, key):
    pig_rows = sorted([r for r in rows if int(r["pig_id"]) == pid], key=lambda x: int(x["day_num"]))
    if not pig_rows:
        return None
    days = [int(r["day_num"]) for r in pig_rows]
    dates = [f"{r['date'][4:6]}/{r['date'][6:8]}" for r in pig_rows]
    trans = [int(r["total_transitions"]) for r in pig_rows]
    ly = [float(r.get("lying_pct", 0) or 0) for r in pig_rows]
    si = [float(r.get("sitting_pct", 0) or 0) for r in pig_rows]
    st = [float(r.get("standing_pct", 0) or 0) for r in pig_rows]
    heat = [r["is_heat_day"] == "True" for r in pig_rows]

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8), sharex=True)
    sow = stall_to_sow.get(pid, f"Stall {pid}")
    title = f"{group_name} — Pig P{pid} ({sow})"
    bar_colors = ["#FF6B6B" if h else "#4A90E2" for h in heat]
    ax1.bar(days, trans, color=bar_colors, edgecolor="black", linewidth=0.5)
    ax1.set_ylabel("Transitions per night")
    ax1.set_title(f"{title}\nNighttime (19:00-08:00) — red bars = heat day")
    ax1.grid(True, alpha=0.3, axis="y")
    for d, t, h, date in zip(days, trans, heat, dates):
        if h:
            ax1.annotate(f"heat\n{date}", xy=(d, t), xytext=(0, 10),
                         textcoords="offset points", ha="center", fontsize=8, color="darkred")
    ax2.stackplot(days, ly, si, st, labels=["Lying", "Sitting", "Standing"],
                  colors=["#3b82f6", "#f97316", "#22c55e"], alpha=0.85)
    for d, h in zip(days, heat):
        if h:
            ax2.axvline(d, color="red", linestyle="--", linewidth=1.5, alpha=0.7)
    ax2.set_ylim(0, 100)
    ax2.set_xlabel("Day # (from group's Day 0)")
    ax2.set_ylabel("Posture %")
    ax2.legend(loc="lower right"); ax2.grid(True, alpha=0.3)
    plt.tight_layout()
    return _save_tmp(fig, key)


def load_predictions_by_pig(predictions_csv):
    """Load raw predictions grouped by pig_id, sorted by timestamp."""
    by_pig = defaultdict(list)
    with open(predictions_csv, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            try:
                dt = datetime.strptime(row["pig_timestamp"], "%Y%m%d-%H-%M-%S")
            except ValueError:
                continue
            by_pig[int(row["pig_id"])].append({
                "datetime": dt,
                "timestamp": row["pig_timestamp"],
                "posture": row["prediction_name"],
                "confidence": float(row["confidence"]),
            })
    for pid in by_pig:
        by_pig[pid].sort(key=lambda x: x["datetime"])
        # Add "change" column
        for i, frame in enumerate(by_pig[pid]):
            if i == 0:
                frame["change"] = 0
            else:
                frame["change"] = 1 if by_pig[pid][i - 1]["posture"] != frame["posture"] else 0
    return by_pig


def generate_report(group):
    trans_rows = load_rows(group["transitions_csv"])
    pig_predictions = load_predictions_by_pig(group["predictions_csv"])
    stall_to_sow = get_stall_to_sow(group["estrus"], group["sow_source"])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    heat_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    # ── Overview ──
    ws = wb.create_sheet("Overview")
    ws.merge_cells("A1:K1")
    ws["A1"] = f'{group["name"]} — Master Report'
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:K2")
    ws["A2"] = "Labels: P{id} = Stall ID / Sow Tag from Lucas's estrus sheet. Nighttime window: 19:00-08:00. Heat = MORNING=20 in estrus."
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    headers = ["Pig ID", "SOW/Tag", "Nights", "Heat Nights", "Total Trans.",
               "Avg/Night", "Heat Avg", "Non-Heat Avg", "Diff", "Lying %", "Standing %"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = thin

    pigs = sorted(set(int(r["pig_id"]) for r in trans_rows))
    r = 5
    for pid in pigs:
        pig_rows = [row for row in trans_rows if int(row["pig_id"]) == pid]
        nights = len(pig_rows)
        heat_rows = [row for row in pig_rows if row["is_heat_day"] == "True"]
        noheat_rows = [row for row in pig_rows if row["is_heat_day"] == "False"]
        total_trans = sum(int(row["total_transitions"]) for row in pig_rows)
        total_frames = sum(int(row["total_frames"]) for row in pig_rows)
        avg = total_trans / nights if nights else 0
        h_avg = sum(int(row["total_transitions"]) for row in heat_rows) / len(heat_rows) if heat_rows else None
        n_avg = sum(int(row["total_transitions"]) for row in noheat_rows) / len(noheat_rows) if noheat_rows else None
        total_lying = sum(int(row.get("lying", 0)) for row in pig_rows)
        total_standing = sum(int(row.get("standing", 0)) for row in pig_rows)
        lying_pct = total_lying / total_frames * 100 if total_frames else 0
        standing_pct = total_standing / total_frames * 100 if total_frames else 0
        vals = [f"P{pid}", stall_to_sow.get(pid, f"Stall {pid}"),
                nights, len(heat_rows), total_trans, round(avg, 1),
                round(h_avg, 1) if h_avg is not None else "-",
                round(n_avg, 1) if n_avg is not None else "-",
                round(h_avg - n_avg, 1) if h_avg is not None and n_avg is not None else "-",
                round(lying_pct, 1), round(standing_pct, 1)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(horizontal="center"); cell.border = thin
        r += 1
    for c in range(1, 12):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # Embed group-wide plots
    plot_row = r + 2
    for title, key_suffix, fn in [
        ("Nightly Transitions (all pigs)", "all_trans", plot_all_pigs_transitions),
        ("Per-pig Posture Breakdown", "all_postures", plot_all_pigs_postures),
    ]:
        ws.cell(row=plot_row, column=1, value=title).font = Font(bold=True, size=12)
        plot_row += 1
        key = f"{group['short']}_{key_suffix}"
        img_path = fn(trans_rows, stall_to_sow, group["name"], key)
        img = XLImage(img_path)
        max_w = 1100
        if img.width > max_w:
            ratio = max_w / img.width
            img.width = max_w
            img.height = int(img.height * ratio)
        ws.add_image(img, f"A{plot_row}")
        plot_row += int(img.height / 18) + 2

    # ── Per-pig sheets ──
    for pid in pigs:
        pig_trans = sorted([row for row in trans_rows if int(row["pig_id"]) == pid],
                           key=lambda x: int(x["day_num"]))
        ws = wb.create_sheet(f"P{pid}")
        sow = stall_to_sow.get(pid, f"Stall {pid}")

        ws.merge_cells("A1:N1")
        ws["A1"] = f"Pig P{pid} ({sow}) — {group['name']}"
        ws["A1"].font = Font(bold=True, size=14)

        # Daily summary table
        ws.cell(row=3, column=1, value="NIGHTLY SUMMARY (19:00-08:00)").font = Font(bold=True, size=12)
        hdrs = ["Date", "Day #", "Heat Day", "Frames", "Transitions",
                "Lying", "Sitting", "Standing", "Lying %", "Sitting %", "Standing %",
                "L->S", "S->L", "L->Si"]
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=5, column=c, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center"); cell.border = thin

        r = 6
        for row in pig_trans:
            is_heat = row["is_heat_day"] == "True"
            date_str = row["date"]
            vals = [f"{date_str[4:6]}/{date_str[6:8]}/{date_str[:4]}", int(row["day_num"]),
                    "YES" if is_heat else "", int(row["total_frames"]), int(row["total_transitions"]),
                    int(row.get("lying", 0)), int(row.get("sitting", 0)), int(row.get("standing", 0)),
                    float(row.get("lying_pct", 0)), float(row.get("sitting_pct", 0)), float(row.get("standing_pct", 0)),
                    int(row.get("lying_to_standing", 0)), int(row.get("standing_to_lying", 0)),
                    int(row.get("lying_to_sitting", 0))]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = Alignment(horizontal="center"); cell.border = thin
                if is_heat:
                    cell.fill = heat_fill
            r += 1
        for c in range(1, 15):
            ws.column_dimensions[get_column_letter(c)].width = 11

        # Embed individual pig plot
        plot_row = r + 2
        ws.cell(row=plot_row, column=1, value="TRANSITIONS + POSTURE TREND").font = Font(bold=True, size=12)
        plot_row += 1
        key = f"{group['short']}_p{pid}"
        img_path = plot_individual_pig(trans_rows, pid, stall_to_sow, group["name"], key)
        if img_path:
            img = XLImage(img_path)
            max_w = 1000
            if img.width > max_w:
                ratio = max_w / img.width
                img.width = max_w
                img.height = int(img.height * ratio)
            ws.add_image(img, f"A{plot_row}")
            plot_row += int(img.height / 18) + 2

        # Raw observations table with "change" column
        raw_start = plot_row
        ws.cell(row=raw_start, column=1, value="RAW OBSERVATIONS (every frame, color-coded)").font = Font(bold=True, size=12)
        raw_hdrs = ["Timestamp", "Posture", "Change", "Confidence"]
        for c, h in enumerate(raw_hdrs, 1):
            cell = ws.cell(row=raw_start + 2, column=c, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center"); cell.border = thin

        rr = raw_start + 3
        for frame in pig_predictions.get(pid, []):
            vals = [frame["timestamp"], frame["posture"], frame["change"], round(frame["confidence"], 3)]
            color = POSTURE_COLORS.get(frame["posture"], "FFFFFF")
            posture_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=rr, column=c, value=v)
                cell.alignment = Alignment(horizontal="center"); cell.border = thin
                if c == 2:  # posture column — color the cell
                    cell.fill = posture_fill
            rr += 1

    out_path = os.path.join(OUTPUT_DIR, f"master_report_{group['short']}.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    tmp = os.path.join(OUTPUT_DIR, ".plot_tmp")
    if os.path.isdir(tmp):
        shutil.rmtree(tmp)

    for g in GROUPS:
        if not os.path.exists(g["transitions_csv"]) or not os.path.exists(g["predictions_csv"]):
            print(f"Missing data files for {g['name']}")
            continue
        generate_report(g)


if __name__ == "__main__":
    main()
